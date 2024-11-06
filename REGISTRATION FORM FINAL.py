from tkinter import * 
from tkinter import messagebox                              #used to display simple pop-up dialogs for user interaction. 
import openpyxl                                             #Excel File Management in GUI Applications
from openpyxl import load_workbook, Workbook                #to load workbook 
from PIL import Image, ImageTk                              #handling and displaying images in Python applications         # Python Imaging Library which provides powerful image processing capabilities
import os                                                   #to work with the operating system and handle file-related 
from tkinter import ttk                                     #used to import the ttk (Themed Tkinter Widgets) #it provides widgets such as label, check button, radio button
from tkcalendar import DateEntry                            # Import DateEntry from tkcalendar
from datetime import datetime
root = Tk()                                                 #creating the main window for your Tkinter GUI application

image_path = r"D:\Desktop\IMG_5543.JPG"
image = Image.open(image_path)
image = image.resize((770, 700), Image.Resampling.LANCZOS)  # Resize the image
bg_image = ImageTk.PhotoImage(image)

# Add the background image
bg_label = Label(root, image=bg_image)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)




file_path = r"D:\Desktop\regis.xlsx"

def initialize_workbook():
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        if "monu" in workbook.sheetnames:
            worksheet = workbook["monu"]
        else:
            worksheet = workbook.create_sheet(title="monu")
            worksheet.append(["Name", "Age", "Phone", "Country", "Date of Birth"])  
    else:
        workbook = Workbook()
        worksheet = workbook.active                       # used to access the currently active sheet in the workbook
        worksheet.title = "monu"
        worksheet.append(["Name", "Age", "Phone", "Country", "Date of Birth"]) 
    return workbook, worksheet

# Initialize workbook and worksheet
a, b = initialize_workbook()                               #a: The workbook object (either loaded from an existing file or newly created)                  b: The worksheet object where the data will be writte

root.title("Registration Form")
root.geometry("800x600")
root.resizable(False, False)

def register(): #performs various checks and, if all the input is valid, saves the data to the Excel file
    name = name_info.get()   #used to retrieve the current value
    age = age_info.get()
    phone = phone_info.get()
    country = country_var.get()
    dob = dob_entry.get()  # Get the date of birth

    if name == "":
        messagebox.showerror("Error", "Please enter your name")
    elif any(x.isdigit() for x in name):
        messagebox.showerror("Error", "Name should contain only alphabets")

    elif age == "":
        messagebox.showerror("Error", "Please enter your age")
    elif not age.isdigit():
        messagebox.showerror("Error", "Age should be an integer only")

    elif phone == "":
        messagebox.showerror("Error", "Please enter your phone number")
    elif not phone.isdigit() or len(phone) != 10:
        messagebox.showerror("Error", "Phone number should be of 10 digits only")

    elif country == "":
        messagebox.showerror("Error", "Please select a country")

    elif dob == "":
        messagebox.showerror("Error", "Please select your date of birth")

    else:
        # Append data to the worksheet
        b.append([name, age, phone, country, dob])  # Added Date of Birth to be saved
        a.save(file_path)
        messagebox.showinfo("Success", "Registration successful")
        clear()

        # Properly close the workbook
        a.close()

def clear():
    name_entry.delete(0, END)   # is used to clear the text in an entry widge
    age_entry.delete(0, END)
    phone_entry.delete(0, END)
    country_combobox.set('')
    dob_entry.set_date(datetime(2000, 1, 1))

Label(root, text="REGISTRATION FORM", font="arial 20 bold underline", bg="red", fg="white").pack(fill="both")
name_info = StringVar() #create a special variable that can hold a string value and be associated with a widget
age_info = StringVar()
phone_info = StringVar()
country_var = StringVar()

# Name
Label(root, text="Enter your Name", font="arial 14").place(x=30, y=70)
name_entry = Entry(root, font="10", bd=4, textvariable=name_info)
name_entry.place(x=400, y=72)

# Age
Label(root, text="Enter your Age", font="arial 14").place(x=30, y=130)
age_entry = Entry(root, font="10", bd=4, textvariable=age_info)
age_entry.place(x=400, y=130)

# Phone Number
Label(root, text="Enter your Mobile Number", font="arial 14").place(x=30, y=190)
phone_entry = Entry(root, font="10", bd=4, textvariable=phone_info)
phone_entry.place(x=400, y=200)

# Country (Dropdown)
Label(root, text="Select your Country", font="arial 14").place(x=30, y=260)

countries = ["USA", "Canada", "UK", "Germany", "France", "India", "Australia", "Brazil", "China", "Japan"]
country_combobox = ttk.Combobox(root, textvariable=country_var, values=countries, state="readonly", font="10", width=28)
country_combobox.place(x=400, y=260)
country_combobox.set("Select Country")  # Default text

# Date of Birth (Calendar)
Label(root, text="Select your Date of Birth", font="arial 14").place(x=30, y=320)
dob_entry = DateEntry(root, width=20, background='darkblue', foreground='white', borderwidth=5, font="10")
dob_entry.place(x=400, y=320)

# Registration button
Button(root, text="Register", font="14", command=register).place(x=400, y=380)

# Clear button
Button(root, text="Clear", font="14", command=clear).place(x=720, y=550)

mainloop()